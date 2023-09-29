import logging

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

import config


class ExcelSaver:
    """Обработчик сохранения Excel файлов."""

    def __init__(
        self,
        ticker,
        direction,
        timeframe_interval,
        s_date,
        u_date,
        all_other_parameters
    ) -> None:
        self.ticker = ticker
        self.direction = direction
        self.timeframe_interval = timeframe_interval
        self.s_date = s_date
        self.u_date = u_date
        self.all_other_parameters = all_other_parameters

    def get_file_path(self):
        """Получение пути сохраняемого файла."""
        file_name = (
            f'{self.direction}_{self.ticker}_{self.timeframe_interval}_'
            f'{self.s_date[:10]}_{self.u_date[:10]}.xlsx'
        )
        return config.RESULTS_DIR + file_name

    @staticmethod
    def adjust_column_width(sheet):
        """
        Автоматически задаем ширину столбцов таблицы Excel
        """
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as ex:
                    logging.error(ex)
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    def write_to_excel(self, file_path):
        """Сохранение Excel файла."""
        headers = [
            'Дата',
            'ticker',
            'open_to_close_trade_duration',

            'Вход',
            'Стоп',
            'Тейк',
            'Результат',
            'Разница',
            'Успешно',
            'Stop',
            'Take',
        ]

        wb = Workbook()
        ws = wb.active

        for col_num, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header

        for row_num, params in enumerate(self.all_other_parameters, start=2):
            for col_num, param in enumerate(params, start=1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = param

        # self.adjust_column_width(ws)

        wb.save(file_path)
        wb.close()

    def save(self):
        """Сохранение файла."""
        result_path = self.get_file_path()
        self.write_to_excel(result_path)
